from typing import Optional, List
from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
import csv
import io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from app.db.database import get_db
from app.schemas.report import ResearcherReportItem, PublicationReportItem, ConferenceReportItem
from app.services.report_service import researcher_report, publication_report, conference_report
from app.core.dependencies import get_current_user, require_roles
from app.models.user import User
from app.utils.constants import UserRole

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get(
    "/researchers",
    response_model=List[ResearcherReportItem],
    dependencies=[Depends(require_roles(UserRole.SYSTEM_ADMIN.value, UserRole.INSTITUTION_ADMIN.value))],
)
def get_researcher_report(
    institution_id: Optional[int] = None,
    department_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return researcher_report(db, institution_id, department_id, current_user)


@router.get(
    "/publications",
    response_model=List[PublicationReportItem],
    dependencies=[Depends(require_roles(UserRole.SYSTEM_ADMIN.value, UserRole.INSTITUTION_ADMIN.value))],
)
def get_publication_report(
    institution_id: Optional[int] = None,
    department_id: Optional[int] = None,
    conference_id: Optional[int] = None,
    status: Optional[str] = None,
    publication_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return publication_report(db, institution_id, department_id, conference_id, status, publication_type, current_user)


@router.get(
    "/conferences",
    response_model=List[ConferenceReportItem],
    dependencies=[Depends(require_roles(UserRole.SYSTEM_ADMIN.value, UserRole.INSTITUTION_ADMIN.value))],
)
def get_conference_report(db: Session = Depends(get_db)):
    return conference_report(db)


@router.get(
    "/publications/export",
    dependencies=[Depends(require_roles(UserRole.SYSTEM_ADMIN.value, UserRole.INSTITUTION_ADMIN.value))],
)
def export_publication_report(
    institution_id: Optional[int] = None,
    department_id: Optional[int] = None,
    conference_id: Optional[int] = None,
    status: Optional[str] = None,
    publication_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    data = publication_report(db, institution_id, department_id, conference_id, status, publication_type, current_user)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        "publication_id", "title", "publication_type", "status",
        "owner_first_name", "owner_last_name", "institution_name",
        "conference_title", "citation_count",
    ])
    writer.writeheader()
    for row in data:
        writer.writerow(row)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=publication_report.csv"},
    )
@router.get(
    "/publications/export/excel",
    dependencies=[Depends(require_roles(UserRole.SYSTEM_ADMIN.value, UserRole.INSTITUTION_ADMIN.value))],
)
def export_publication_report_excel(
    institution_id: Optional[int] = None,
    department_id: Optional[int] = None,
    conference_id: Optional[int] = None,
    status: Optional[str] = None,
    publication_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    data = publication_report(db, institution_id, department_id, conference_id, status, publication_type, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Publication Report"

    headers = ["ID", "Title", "Type", "Status", "Author", "Institution", "Conference", "Citations"]
    ws.append(headers)

    header_fill = PatternFill(start_color="0F1B2E", end_color="0F1B2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in data:
        ws.append([
            row["publication_id"],
            row["title"],
            row["publication_type"],
            row["status"],
            f"{row['owner_first_name']} {row['owner_last_name']}",
            row["institution_name"],
            row["conference_title"] or "-",
            row["citation_count"],
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 45)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=publication_report.xlsx"},
    )


@router.get(
    "/publications/export/pdf",
    dependencies=[Depends(require_roles(UserRole.SYSTEM_ADMIN.value, UserRole.INSTITUTION_ADMIN.value))],
)
def export_publication_report_pdf(
    institution_id: Optional[int] = None,
    department_id: Optional[int] = None,
    conference_id: Optional[int] = None,
    status: Optional[str] = None,
    publication_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    data = publication_report(db, institution_id, department_id, conference_id, status, publication_type, current_user)

    stream = io.BytesIO()
    doc = SimpleDocTemplate(stream, pagesize=landscape(A4))

    table_data = [["ID", "Title", "Type", "Status", "Author", "Institution", "Conference", "Citations"]]
    for row in data:
        table_data.append([
            str(row["publication_id"]),
            row["title"][:40],
            row["publication_type"],
            row["status"],
            f"{row['owner_first_name']} {row['owner_last_name']}",
            row["institution_name"],
            (row["conference_title"] or "-")[:25],
            str(row["citation_count"]),
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F1B2E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F6F2")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    doc.build([table])
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=publication_report.pdf"},
    )